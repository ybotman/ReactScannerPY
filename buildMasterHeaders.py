#!/usr/bin/env python3
import os
import sys
import logging
from pathlib import Path
import fnmatch
from openpyxl import load_workbook, Workbook

# Set logging to DEBUG or INFO as needed
logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')

def load_worksheet(excel_path, sheet_name):
    wb = load_workbook(excel_path)
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    return wb, wb[sheet_name]

def is_git_initialized(folder_path: Path) -> bool:
    git_exists = (folder_path / '.git').exists()
    logging.debug(f"Checking git init in {folder_path}: {git_exists}")
    return git_exists

def load_gitignore_patterns(folder_path: Path):
    """Load .gitignore patterns if available."""
    gitignore_path = folder_path / '.gitignore'
    patterns = []
    if gitignore_path.exists():
        logging.info(f"Loading .gitignore from {gitignore_path}")
        with gitignore_path.open('r', encoding='utf-8') as f:
            for line in f:
                line = line.strip()
                if not line or line.startswith('#'):
                    continue
                patterns.append(line)
    else:
        logging.info(f"No .gitignore found at {folder_path}, no patterns loaded.")
    return patterns

def path_matches_patterns(path_str: str, patterns):
    """Check if a path matches any gitignore patterns."""
    # We'll apply fnmatch for each pattern.
    # If a pattern ends with '/', it means a directory.
    # We'll just use fnmatch to check either way.
    for p in patterns:
        if fnmatch.fnmatch(path_str, p):
            return True
    return False

def should_skip_dir(rel_dir_str, patterns):
    """Check if this directory should be skipped based on .gitignore patterns."""
    # Directories like 'node_modules/' in .gitignore should skip the dir and everything inside.
    # If any pattern matches the directory name, skip it.
    # We'll try it as-is and also with trailing slash.
    if path_matches_patterns(rel_dir_str + '/', patterns):
        return True
    return False

def get_depth(root: Path, file_path: Path) -> int:
    # Depth: number of directories between root and file (not counting the file itself)
    return len(file_path.relative_to(root).parts) - 1

def scan_folder(root: Path, patterns):
    """Recursively find .js and .css files, ignoring hardcoded patterns and .gitignore rules."""
    results = []
    logging.info(f"Scanning folder: {root}")

    # Hardcoded directories to ignore
    hardcoded_ignores = {".next", "node_modules", "archive"}

    for dirpath, dirnames, filenames in os.walk(root):
        rel_dir = Path(dirpath).relative_to(root)

        # Check for hardcoded ignores
        if any(part in hardcoded_ignores for part in rel_dir.parts):
            logging.debug(f"Skipping hardcoded directory: {rel_dir}")
            dirnames[:] = []  # Clear subdirs to skip descending
            continue

        # Check if this directory should be skipped based on .gitignore
        if str(rel_dir) != '.' and should_skip_dir(str(rel_dir), patterns):
            logging.debug(f"Skipping directory: {rel_dir} due to .gitignore")
            dirnames[:] = []  # Clear subdirs to skip descending
            continue

        # Scan files in the directory
        for filename in filenames:
            fpath = Path(dirpath) / filename
            rel_file = fpath.relative_to(root)

            # Check file patterns
            if fpath.suffix in ['.js', '.css']:
                # Check if file matches .gitignore
                if path_matches_patterns(str(rel_file), patterns):
                    logging.debug(f"Skipping file {rel_file} due to .gitignore")
                    continue
                logging.debug(f"Found file {rel_file}")
                results.append(fpath)

    return results

def main():
    # Constants
    excel_path = Path("App_React_PathFiles_selector.xlsx")
    codefolders_sheet_name = "CodeFolders"
    code_sheet_name = "CODE"

    # Load workbook and sheets
    if not excel_path.exists():
        logging.error("Excel file not found: %s", excel_path)
        sys.exit(1)
    wb, codefolders_sheet = load_worksheet(excel_path, codefolders_sheet_name)
    if codefolders_sheet_name not in wb.sheetnames:
        logging.error("No CodeFolders tab found in Excel file.")
        sys.exit(1)
    if code_sheet_name not in wb.sheetnames:
        wb.create_sheet(code_sheet_name)
    code_sheet = wb[code_sheet_name]

    # Read CodeFolders tab
    headers = [cell.value for cell in codefolders_sheet[1]]
    folder_col = None
    wantsan_col = None
    for idx, h in enumerate(headers, start=1):
        if h == "Folder":
            folder_col = idx
        elif h == "WantScan":
            wantsan_col = idx

    if folder_col is None or wantsan_col is None:
        logging.error("CodeFolders sheet must have Folder and WantScan columns.")
        sys.exit(1)

    codefolders = []
    for row in codefolders_sheet.iter_rows(min_row=2, values_only=True):
        folder_val = row[folder_col-1]
        wantscan_val = row[wantsan_col-1]
        if folder_val and isinstance(wantscan_val, bool):
            codefolders.append((folder_val, wantscan_val))

    # Validate git repos and scan
    all_scanned_records = []
    for folder_path_str, want_scan in codefolders:
        if not want_scan:
            logging.info(f"Skipping folder {folder_path_str} because WantScan=False")
            continue
        folder_path = Path(folder_path_str)
        if not folder_path.exists() or not folder_path.is_dir():
            logging.warning("Folder does not exist or not a directory: %s", folder_path)
            continue
        if not is_git_initialized(folder_path):
            logging.warning("Folder not git-initialized: %s", folder_path)
            continue

        # Load .gitignore
        patterns = load_gitignore_patterns(folder_path)
        scanned_files = scan_folder(folder_path, patterns)
        for f in scanned_files:
            depth = get_depth(folder_path, f)
            all_scanned_records.append((str(folder_path.resolve()), str(f.resolve()), f.name, depth))

    # Update CODE sheet
    logging.info("Updating CODE sheet with scanned files.")
    code_headers = [cell.value for cell in code_sheet[1]] if code_sheet.max_row > 0 else []
    if not code_headers or any(h not in code_headers for h in ["Path", "File", "Depth", "WantDoc"]):
        code_sheet.delete_rows(1, code_sheet.max_row)
        code_sheet.append(["Path", "File", "Depth", "WantDoc"])
        code_headers = ["Path", "File", "Depth", "WantDoc"]

    path_idx = code_headers.index("Path") + 1
    file_idx = code_headers.index("File") + 1
    depth_idx = code_headers.index("Depth") + 1
    wantdoc_idx = code_headers.index("WantDoc") + 1

    existing_rows = {}
    for row in code_sheet.iter_rows(min_row=2, values_only=False):
        p = row[path_idx-1].value
        f = row[file_idx-1].value
        d = row[depth_idx-1].value
        w = row[wantdoc_idx-1].value
        if p:
            existing_rows[p] = (f, d, w)

    # Clear CODE rows except header
    code_sheet.delete_rows(2, code_sheet.max_row)

    new_paths = set(r[1] for r in all_scanned_records)
    for full_path in sorted(new_paths):
        old = existing_rows.get(full_path)
        record = next((r for r in all_scanned_records if r[1] == full_path), None)
        if record:
            _, _, new_fname, new_depth = record
            if old:
                # Keep old WantDoc
                _, _, old_w = old
                code_sheet.append([full_path, new_fname, new_depth, old_w])
            else:
                # New file, WantDoc = False
                code_sheet.append([full_path, new_fname, new_depth, False])
        else:
            # Should not happen, since we got it from new_paths
            pass

    wb.save(excel_path)
    logging.info("Excel file updated successfully.")

if __name__ == "__main__":
    main()