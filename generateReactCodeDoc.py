#!/usr/bin/env python3
import os
import sys
import logging
from pathlib import Path
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format='%(asctime)s [%(levelname)s] %(message)s')

def main():
    excel_path = Path("App_React_PathFiles_selector.xlsx")
    code_sheet_name = "CODE"
    codefolders_sheet_name = "CodeFolders"

    if not excel_path.exists():
        logging.error("Excel file not found: %s", excel_path)
        sys.exit(1)
    wb = load_workbook(excel_path)
    if code_sheet_name not in wb.sheetnames:
        logging.error("CODE sheet not found in Excel file.")
        sys.exit(1)
    if codefolders_sheet_name not in wb.sheetnames:
        logging.error("CodeFolders sheet not found in Excel file.")
        sys.exit(1)

    code_sheet = wb[code_sheet_name]
    codefolders_sheet = wb[codefolders_sheet_name]

    # Load codefolders to help determine relative paths
    cf_headers = [c.value for c in codefolders_sheet[1]]
    if "Folder" not in cf_headers:
        logging.error("CodeFolders sheet missing 'Folder' column.")
        sys.exit(1)
    folder_col = cf_headers.index("Folder") + 1
    codefolders = []
    for row in codefolders_sheet.iter_rows(min_row=2, values_only=True):
        fpath = row[folder_col-1]
        if fpath:
            codefolders.append(str(Path(fpath).resolve()))

    # Expect: Path, File, Depth, WantDoc
    headers = [c.value for c in code_sheet[1]]
    try:
        path_col = headers.index("Path") + 1
        file_col = headers.index("File") + 1
        depth_col = headers.index("Depth") + 1
        wantdoc_col = headers.index("WantDoc") + 1
    except ValueError:
        logging.error("CODE sheet missing required columns.")
        sys.exit(1)

    entries = []
    for row in code_sheet.iter_rows(min_row=2, values_only=True):
        p, f, d, w = row[path_col-1], row[file_col-1], row[depth_col-1], row[wantdoc_col-1]
        if p and f and d is not None and w is not None:
            if w is True:
                entries.append((p, f, d))

    entries.sort(key=lambda x: x[0])

    desktop_path = Path.home() / "Desktop" / "AppReactDocument.txt"

    # We need to print a relative folder marker.
    # For each file:
    # 1. Find which CodeFolder is a prefix of this path. Choose the longest matching prefix.
    # 2. relative_path = file_path.relative_to(parent_of_that_codefolder's_parent)
    # 3. Print "——>>> FOLDER : ../{that_relative_path_without_filename} <<<——" 
    #    We'll add '../' at the start to mimic the requested style.

    def find_best_codefolder(file_path: Path, codefolders):
        best_match = None
        file_str = str(file_path)
        for cf in codefolders:
            if file_str.startswith(cf):
                # Pick the longest match
                if best_match is None or len(cf) > len(best_match):
                    best_match = cf
        return best_match

    with desktop_path.open('w', encoding='utf-8') as out:
        for (full_path, fname, depth) in entries:
            file_path = Path(full_path)
            parent_dir = file_path.parent

            # Determine relative folder marker
            folder_marker = str(parent_dir)
            best_cf = find_best_codefolder(file_path, codefolders)
            if best_cf:
                # Make a relative path from parent_of_best_cf's parent
                best_cf_path = Path(best_cf)
                parent_of_cf = best_cf_path.parent
                # relative from parent_of_cf
                rel = file_path.parent.relative_to(parent_of_cf)
                # Prepend '../'
                folder_marker = "../" + str(rel)

            if depth <= 4:
                out.write(f"——>>> FOLDER : {folder_marker} <<<——\n")
            out.write("——>>> users code <<<——\n")

            try:
                with file_path.open('r', encoding='utf-8') as f:
                    file_content = f.read()
                out.write(file_content)
            except Exception as e:
                logging.error("Error reading file %s: %s", file_path, e)
                out.write(f"\n[Error reading file: {e}]\n")

            # Add 5 blank lines
            out.write("\n" * 5)

    logging.info(f"AppReactDocument.txt created on Desktop at {desktop_path}")

if __name__ == "__main__":
    main()